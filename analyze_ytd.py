import openpyxl, json, os, traceback

try:
    base = r'C:\Users\dposavac\CLionProjects\Excel_transfer'
    fpath = os.path.join(base, 'YTD_01_2026.xlsm')
    print("File exists:", os.path.exists(fpath))
    print("File size:", os.path.getsize(fpath))
    
    wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
    print('Sheet names:', wb.sheetnames)
    
    with open(os.path.join(base, 'JSON', 'mappings_sap_ytd.json'), 'r') as f:
        mappings = json.load(f)
    
    march = next((m for m in mappings if m.get('month','').lower() == 'march'), None)
    src_sheet = march.get('source_sheet', 'Report')
    print('source_sheet=%s, source_column=%s' % (src_sheet, march.get('source_column','C')))
    
    sheet_name = None
    for sn in wb.sheetnames:
        if sn.lower() == src_sheet.lower():
            sheet_name = sn
            break
    if not sheet_name:
        print('Sheet "%s" NOT FOUND' % src_sheet)
        for sn in wb.sheetnames:
            if '01_2026' in sn or '03_2026' in sn:
                sheet_name = sn
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
    
    print('Using sheet:', sheet_name)
    ws = wb[sheet_name]
    
    rowMap = march.get('rowMap', {})
    print('rowMap has %d entries' % len(rowMap))
    print()
    
    print('destRow -> srcRow : value')
    for dest_str, src in list(rowMap.items())[:25]:
        if isinstance(src, int): srs = [src]
        elif isinstance(src, list): srs = src
        elif isinstance(src, dict) and 'sum' in src: srs = src['sum']
        else: srs = [src]
        for sr in srs:
            if sr == 0:
                print('  dest %4s -> src %5d : zero' % (dest_str, sr))
                continue
            c = ws.cell(row=abs(sr), column=3)
            print('  dest %4s -> src %5d : val=%r type=%s' % (dest_str, abs(sr), c.value, type(c.value).__name__))
    
    print()
    print('Key rows:')
    for r in [9, 35, 56, 466, 587, 907, 1429, 2582]:
        c = ws.cell(row=r, column=3)
        print('  Row %5d Col C: val=%r type=%s' % (r, c.value, type(c.value).__name__))
    
    print('Max row:', ws.max_row)
    wb.close()

except Exception as e:
    traceback.print_exc()